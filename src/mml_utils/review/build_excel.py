import copy
import csv
import datetime
import pathlib
import random

from loguru import logger


def _dump_current_sample(outpath, feature_name, dt, note_ids):
    """Dump previous sample"""
    with open(_get_sample_path(outpath) / f'sample.{feature_name}.{dt}.txt', 'w') as out:
        out.write('\n'.join(str(x) for x in note_ids))


def _get_sample_path(outpath: pathlib.Path):
    path = outpath / '..' / 'sample'
    path.mkdir(exist_ok=True)
    return path


def _get_previous_samples(outpath, feature_name):
    """Get previous sampled note ids to avoid sampling same note id multiple times."""
    results = set()
    cnt = 0  # can't use len() due to adding both string and int versions
    # get all previous samples from parent directory (i.e., the original outpath)
    for file in _get_sample_path(outpath).glob(f'sample.{feature_name}.*.txt'):
        with open(file) as fh:
            for note_id in fh:
                nid = note_id.strip()
                if not nid or nid in results:
                    continue
                results.add(nid)
                cnt += 1
                try:
                    results.add(int(nid))  # add int version too, if possible
                except ValueError:
                    continue
    logger.info(f'Ignoring {cnt} previously sampled ids for {feature_name}.')
    return results


def compile_to_excel(outpath: pathlib.Path, note_ids, text_encoding='utf8', sample_size=50,
                     metadata_file=None, build_csv=False, sample_snippets_per_note=None):
    """

    :param build_csv:
    :param sample_snippets_per_note: sample up to this many snippets for each note (this was set to 1 for an IRR pull)
    :param note_ids:
    :param outpath:
    :param text_encoding:
    :param sample_size: number of notes to sample
    :param metadata_file: csv file with:
        HEADER = note_id, studyid, date, etc.
        VALUES = 1, A2E, 05/06/2011, etc.
    :return:
    """
    # e.g., HEADER = note_id, studyid, date, etc.
    # e.g., VALUES = 1, A2E, 05/06/2011
    metadata_lkp = {}
    new_fields = []
    if metadata_file:
        with open(metadata_file, newline='') as fh:
            reader = csv.DictReader(fh)
            new_fields = set(reader.fieldnames) - {'note_id'}
            for row in reader:
                note_id = row['note_id']
                del row['note_id']
                metadata_lkp[note_id] = row

    # select random sample
    new_note_ids = {}
    dt = datetime.datetime.now().strftime('%Y%m%d%H%M%S')
    for feature_name in note_ids:
        if sample_size:
            prev_samples = _get_previous_samples(outpath, feature_name)
            curr_pop = set(note_ids[feature_name]) - prev_samples
            new_note_ids[feature_name] = set(
                random.sample(curr_pop, min([sample_size, len(curr_pop)]))
            )
            _dump_current_sample(outpath, feature_name, dt, new_note_ids[feature_name])
        else:
            new_note_ids[feature_name] = None

    has_excel = False
    try:
        import openpyxl

        has_excel = True
    except ImportError:
        logger.warning(f'Openpyxl not installed: not building excel workbook, just CSV files.')
        logger.info(f'To build Excel workbook, install openpyxl: `pip install openpyxl` and re-run.')

    config = None
    if has_excel:
        config = _build_excel_review_set(outpath, new_fields, new_note_ids, metadata_lkp, sample_size, text_encoding,
                                         sample_snippets_per_note)
    if not has_excel or build_csv:
        logger.info(f'Building CSV file output.')
        _build_csv_review_set(outpath, new_fields, new_note_ids, metadata_lkp, sample_size, text_encoding,
                              sample_snippets_per_note, config)


def _build_csv_review_set(outpath, new_fields, note_ids, metadata_lkp, sample_size, text_encoding,
                          sample_snippets_per_note=None, config=None):
    for file in outpath.parent.glob('*.review.csv'):
        feature_name = file.name.split('.')[0]
        sample_note_ids = note_ids[feature_name] if note_ids[feature_name] else None
        curr_note_id_records = []
        curr_note_id = None
        with open(file, newline='', encoding=text_encoding) as fh:
            reader = csv.DictReader(fh)
            with open(outpath / f'{feature_name}.sample{sample_size}.csv', 'w',
                      newline='', encoding=text_encoding) as out:
                writer = csv.DictWriter(out, fieldnames=reader.fieldnames[:7] + new_fields + reader.fieldnames[7:])
                writer.writeheader()
                for row in reader:
                    new_note_id = row['note_id']
                    if sample_note_ids and new_note_id not in sample_note_ids:
                        continue

                    if curr_note_id != new_note_id:
                        _writerows(writer, curr_note_id, curr_note_id_records, sample_snippets_per_note, config)
                        curr_note_id = new_note_id
                        curr_note_id_records = []
                    curr_note_id_records.append({**row, **metadata_lkp.get(new_note_id, dict())})

                # final fencepost
                _writerows(writer, curr_note_id, curr_note_id_records, sample_snippets_per_note, config)


def _writerows(writer, curr_note_id, curr_note_id_records, sample_snippets_per_note, config):
    if curr_note_id_records:  # nothing here first round
        if sample_snippets_per_note and config:
            writer.writerows(
                [curr_note_id_records[i] for i in config.get('sample-snippet')[curr_note_id]]
            )
        elif sample_snippets_per_note:  # nothing yet sampled
            length = len(curr_note_id_records)
            indices = random.sample(range(length), min(length, sample_snippets_per_note))
            writer.writerows([curr_note_id_records[i] for i in indices])
        else:  # include all snippets
            writer.writerows(curr_note_id_records)


def _build_excel_review_set(outpath, new_fields, note_ids, metadata_lkp, sample_size, text_encoding,
                            sample_snippets_per_note=None):
    """

    :param outpath:
    :param new_fields:
    :param note_ids:
    :param sample_size:
    :param text_encoding:
    :return:
    """
    from openpyxl import Workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    # config: to pass info to CSV files if those are being built
    config = {}
    if sample_snippets_per_note:
        config['sample-snippet'] = {}

    # column data
    pre_columns = ['id', 'note_id', 'start', 'end', 'length', 'negation', 'type']
    post_columns = ['precontext',
                    'keyword',
                    'postcontext',
                    'Is this term referring to the feature?',
                    'Does the patient have this feature?',
                    'Optional Comments',
                    'fullcontext',
                    ]
    column_widths = {**{
        'id': 12,
        'note_id': 12,
        'start': 12,
        'end': 12,
        'length': 12,
        'negation': 12,
        'type': 12,
    }, **{
        field: 12 for field in new_fields
    }, **{
        'precontext': 40,
        'keyword': 24,
        'postcontext': 40,
        'Is this term referring to the feature?': 24,
        'Does the patient have this feature?': 24,
        'Optional Comments': 24,
        'fullcontext': 60,
    }}
    column_alignments = [
                            [True, 'right', 'center'],
                            [True, 'right', 'center'],
                            [True, 'right', 'center'],
                            [True, 'right', 'center'],
                            [True, 'right', 'center'],
                            [True, 'right', 'center'],
                            [True, 'right', 'center'],
                        ] + [
                            [True, 'right', 'center'] for _ in new_fields
                        ] + [
                            [True, 'right', 'center'],
                            [True, 'center', 'center'],
                            [True, 'left', 'center'],
                            [True, 'center', 'center'],
                            [True, 'center', 'center'],
                            [True, 'left', 'center'],
                            [True, 'left', 'center'],
                        ]
    wb = Workbook()
    # Add a default style with striped rows and banded columns
    style = TableStyleInfo(name=f'TableStyleMedium9', showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    for file in outpath.parent.glob('*.review.csv'):
        feature_name = file.name.split('.')[0]
        if feature_name not in note_ids:
            logger.warning(f'No notes found for feature {feature_name}.')
            continue
        sample_note_ids = note_ids[feature_name]
        max_row_index = 1  # 1-indexed
        ws = wb.create_sheet(title=f'{feature_name}')
        with open(file, newline='', encoding=text_encoding) as fh:
            reader = csv.DictReader(fh)
            max_col_letter = get_column_letter(len(column_alignments))
            ws.append(list(column_widths.keys()))
            prev_note_id = None
            prev_note_id_records = []
            for row in reader:
                curr_note_id = row['note_id']
                if sample_note_ids and curr_note_id not in sample_note_ids:
                    continue
                if prev_note_id != curr_note_id:
                    max_row_index = add_to_worksheet(ws, max_row_index, prev_note_id, prev_note_id_records,
                                                     sample_snippets_per_note, config)
                    prev_note_id_records = []
                    prev_note_id = curr_note_id

                if curr_note_id in metadata_lkp:
                    md = list(metadata_lkp[curr_note_id].values())
                else:
                    md = [''] * len(new_fields)
                prev_note_id_records.append(
                    [row[k] for k in pre_columns]
                    + md
                    + [row.get(k, '') for k in post_columns]
                )
            # fencepost
            max_row_index = add_to_worksheet(ws, max_row_index, prev_note_id, prev_note_id_records,
                                             sample_snippets_per_note, config)

            # build/format table
            table = Table(
                displayName=feature_name,
                ref=f'A1:{max_col_letter}{max_row_index}',
                tableStyleInfo=style,
            )
            ws.add_table(table)
            for i, width in enumerate(column_widths.values(), start=1):
                ws.column_dimensions[get_column_letter(i)].width = width
            for sheet_row in ws.iter_rows():
                for i, cell in enumerate(sheet_row):
                    alignment = copy.copy(cell.alignment)
                    alignment.wrapText = column_alignments[i][0]
                    alignment.horizontal = column_alignments[i][1]
                    alignment.vertical = column_alignments[i][2]
                    cell.alignment = alignment
    wb.remove(wb[wb.sheetnames[0]])  # remove default sheet
    if len(wb.sheetnames) > 0:
        wb.save(outpath / f'features.review.sample{sample_size}.xlsx')
    else:
        logger.error(f'No features contain any samples for review: not outputting Excel file.')
        logger.warning(f'Perhaps the *.review.csv should be moved out of the review_DATE folder into the target_dir?')
    return config  # pass configuration to CSV files generated


def add_to_worksheet(ws, max_row_index, prev_note_id, prev_note_id_records, sample_snippets_per_note, config):
    if prev_note_id_records:
        if sample_snippets_per_note is not None:
            length = len(prev_note_id_records)
            indices = random.sample(range(length), min(length, sample_snippets_per_note))
            config['sample-snippet'][prev_note_id] = indices
            for i in indices:
                ws.append(prev_note_id_records[i])
                max_row_index += 1
        else:
            for record in prev_note_id_records:
                ws.append(record)
                max_row_index += 1
    return max_row_index
